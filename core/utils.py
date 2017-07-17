# utils.py
import csv
import codecs
import cStringIO
import time
from datetime import datetime
from pytz import timezone
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.db import IntegrityError
from django.core.urlresolvers import reverse
from django.contrib.auth import logout
from django.contrib.auth.models import User
from django.shortcuts import redirect
from django.conf import settings

from social.exceptions import SocialAuthBaseException, AuthException, AuthForbidden

from courses.models import Course
from lessons.models import Lesson

from .models import ActivityLog, AttendanceTracker

tz = timezone(settings.TIME_ZONE)


class GGVExcelWriter:


    wb = None

    def __init__(self):
        self.wb = Workbook()

    def write_row(self, ws, data=[]):
        ws.append(data)

    def save(self, f):
        self.wb.save(f)

    # def write_row_1(self, ws, u):
    #     ws.append(['GEDid', 'users name', 'users email', datetime.now()])
    #     ws.append([])  # Blank row
    #     for i in self.USER_INFO_CELLS:
    #         ws[i].font = Font(size=16, name='Arial')

    # def setup_data_cols(self, ws):
    #     for col_num in xrange(len(self.DATA_COLS)):
    #         offset = col_num+1
    #         cell = ws.cell(row=3, column=offset)
    #         cell.value = self.DATA_COLS[col_num][1]
    #         cell.font = Font(size=14, name='Arial')
    #         # set column width
    #         ws.column_dimensions[self.DATA_COLS[col_num][0]].width = self.DATA_COLS[col_num][2]


class UnicodeWriter:
    """
    A CSV writer which will write rows to CSV file "f",
    which is encoded in the given encoding. Needed for unicode in csv
    writer.
    """

    def __init__(self, f, dialect=csv.excel, encoding="utf-8", **kwds):
        # Redirect output to a queue
        self.queue = cStringIO.StringIO()
        self.writer = csv.writer(self.queue, dialect=dialect, **kwds)
        self.stream = f
        self.encoder = codecs.getincrementalencoder(encoding)()

    def writerow(self, row):
        self.writer.writerow([s.encode("utf-8") for s in row])
        # Fetch UTF-8 output from the queue ...
        data = self.queue.getvalue()
        data = data.decode("utf-8")
        # ... and reencode it into the target encoding
        data = self.encoder.encode(data)
        # write to the target stream
        self.stream.write(data)
        # empty queue
        self.queue.truncate(0)

    def writerows(self, rows):
        for row in rows:
            self.writerow(row)


class GGVAuthForbidden(SocialAuthBaseException):
    """Auth process exception."""
    def __init__(self, backend, response, gaccount, *args, **kwargs):
        self.backend = backend
        self.gaccount = gaccount
        self.response = response
        super(GGVAuthForbidden, self).__init__(*args, **kwargs)

    def __str__(self):
        m = self.gaccount
        return m


def logout_clean(request):
    logout(request)
    return redirect('https://accounts.google.com/Logout?&continue=https://www.google.com')

def auth_allowed(response, details, *args, **kwargs):
    """
    Return the user object if authenticated email matches a user email.
    Implies that allowed users must be created in system beforehand, with an
    email that matches the gmail account used to authenticate.
    """
    try:
        registration_email_str = details.get('email')

        # Use case insensitive lookup on the email just to be sure.       
        ggv_user = User.objects.get(username__iexact=registration_email_str)
        return ggv_user

    except Exception as e:
        # print e
        return None

def ggv_auth_allowed(backend, details, response, *args, **kwargs):
    """
    If auth_allowed returns a user object, set the user variable for the pipeline.
    A valid user variable is processed to determine if a social (google) association needs
    to be created. See ggv_social for the next op in the pipeline.
    """
    ggv_user = auth_allowed(response, details)
    if not ggv_user:
        raise GGVAuthForbidden(backend, response, details['email'])
    else:
        return {'user': ggv_user}

def ggv_social_user(backend, uid, user=None, *args, **kwargs):
    """
    Previous pipeline op, ggv_auth_allowed, should prevent a null user but
    checking here, again, just in case.
    The effect of this op is that a social association is set. If social association
    is None, the pipeline will create a new social association to the ggv user object.
    Subsequent (non overridden pipeline ops) will process as designed, based on the social and user
    variables being initialized with values.
    """
    if user:
        provider = backend.name
        social = backend.strategy.storage.user.get_social_auth(provider, uid)
        # print 'ggv_social_user=>', social
        if not social:  # user has not logged in previously (e.g., no social auth obj exists) -- make their account active.
            user.is_active = True
            user.save()
    else:
        # print 'ggv_social_user=>', user
        raise AuthForbidden(backend)

    return {'social': social,
            'user': user,
            'is_new': user is None,
            'new_association': False}

def get_daily_log_times(user=None, course=None, exclusions=[]):
    """
    Method should return a list containing dicts described as follows:
    [
        {'day': date, 'duration': seconds, 'events': [event_dict_list]}, 
        ... 
        {'day': date, 'duration': seconds, 'events': [event_dict_list]}
    ]
    """
    
    a = ActivityLog.objects.filter(user__id=user.id)
    acts = []
    act_log = {}

    act_secs = 0
    act_day = None
    for i in range(len(a)-1):
        curr_act_ts = a[i].timestamp.astimezone(tz).strftime('%b-%d-%Y')
        next_act_ts = a[i+1].timestamp.astimezone(tz).strftime('%b-%d-%Y')

        # new date encountered in list, save and reset to a new act_log
        if act_day != curr_act_ts:
            if act_log:
                act_log['duration'] = '%s hours %s minutes' % (act_log['duration']/3600, (act_log['duration']%3600)/60)
                acts.append(act_log)  # Before resetting, append previous daily log dict if exists
            act_day = curr_act_ts
            act_log = {}  
            act_log['day'] = act_day
            act_log['duration'] = 0  
            act_log['events'] = []          
            
        if a[i].action != 'login' and curr_act_ts == next_act_ts: # don't compute duration, login indicates entry event.
            act_log['duration'] = act_log['duration'] + (a[i].timestamp.astimezone(tz)-a[i+1].timestamp.astimezone(tz)).seconds
        
        event_dict = a[i].as_dict(course, exclusions)
        if event_dict:
            act_log['events'].append(event_dict)

    # Save tail end of activity list
    if act_log:
        act_log['duration'] = '%s hours %s minutes' % (act_log['duration']/3600, (act_log['duration']%3600)/60)
        acts.append(act_log)

    return acts

def get_daily_log_times_v2 (user=None, course=None, exclusions=[]):
    """
    Method should return a list containing an ordered dict described as follows:
    [{
        '2016-04-16': [date, duration, [event_dict_list]], 
        ... 
        '2016-01-01': [date, duration, [event_dict_list]]
    }]
    """
    acts = OrderedDict()    
    
    act_list = user.activitylog.all()
    att_list = user.attendance.all()

    for event in act_list:
        curr = event.timestamp_tz()
        curr_datestr = curr.strftime('%Y-%m-%d')
        try:

            acts[curr_datestr][2].append(event.as_dict(course, exclusions))

        except KeyError as e:
            # print curr_datestr

            att_record = att_list.filter(datestr=curr_datestr)
            dur = ""
            if att_record:
                # print curr_datestr, att_record
                secs = att_record[0].duration_in_secs
                # dur = '%s:%s' % (secs/3600, secs%3600/60)
                dur = time.strftime('%-H:%M', time.gmtime(secs))
            
            acts[curr_datestr] = [curr, dur, [event.as_dict(course, exclusions)]]
    return acts

def elapsed_daily_activity(user=None):
    activity = user.activitylog.all()
    log = {} # day: [[a1, a2, ...], elapsed]

    for curr_act in activity:
        t = curr_act.timestamp_tz()
        day = t.strftime('%Y-%m-%d')
        try:
            log[day][0].append(curr_act)
            next_act = log[day][0][ len(log[day][0])-2 ] # previous read
            
            seq = (curr_act.action, next_act.action)
            if next_act.action == 'login' or curr_act.action == 'login':
                pass
            elif seq == ('logout', 'login'):   #curr_act.action == 'logout' or next_act.action == 'login':
                pass
            elif seq == ('login', 'logout'):  # curr_act.action == 'login' and next_act.action == 'logout':
                pass
            else:
                delta = next_act.timestamp_tz() - t
                """ patch auto logout glitch requires this in order to limit over calculation """
                if delta.seconds > 3600:
                    log[day][1] += 3600
                    # print 'Exceeded threshold==> ', delta.seconds
                else:
                    log[day][1] += delta.seconds
                """ END patch """
                
            # print t, curr_act.action, next_act.action, 'ELAPSED:', log[day][1]/60, 'mins', log[day][1]%60, 'secs' 

            
        except KeyError:
            log[day] = [[curr_act], 0]
            # print t, curr_act.action, 0, 'LAST ACTION ON THIS DAY', 'ELAPSED:', 0   

    return log

def elapsed_time_per_event(user=None):
    activity = user.activitylog.all()
    log = []
    subject_time = {}

    for curr_act in range(1, len(activity)):
        act = activity[curr_act]
        t = act.timestamp_tz()
        day = t.strftime('%Y-%m-%d')

        event = [day, act.message_detail]
        try:
            next_act =  activity[curr_act-1]
            
            seq = (act.action, next_act.action)
            if next_act.action == 'login' or act.action == 'login' or next_act.timestamp_tz().strftime('%Y-%m-%d') != day:
                pass
            elif seq == ('logout', 'login'):   #curr_act.action == 'logout' or next_act.action == 'login':
                pass
            elif seq == ('login', 'logout'):  # curr_act.action == 'login' and next_act.action == 'logout':
                pass
            else:
                delta = next_act.timestamp_tz() - t
                
                """ patch auto logout glitch requires this in order to limit over calculation """
                if delta.seconds > 3600:
                    event.append(3600)
                    try:
                        subject_time[act.message_detail] += 3600
                    except:
                        subject_time[act.message_detail] = 3600
                else:
                    event.append(delta.seconds)
                    try:
                        subject_time[act.message_detail] += delta.seconds
                    except:
                        subject_time[act.message_detail] = delta.seconds

                """ END patch """
        except Exception as e:
            print e 

        log.append(event)

    return log, subject_time

def populate_attendance_duration_field(user=None):
    elapsed = elapsed_daily_activity(user)
    for i, j in elapsed.items():
        try:
            a = user.attendance.get(datestr=i)
            a.duration_in_secs = j[1]
            a.save()
        except ObjectDoesNotExist as e:
            a = AttendanceTracker( 
                    user=user, 
                    datestamp=j[0][0].timestamp, 
                    datestr=j[0][0].timestamp_tz().strftime('%Y-%m-%d'),
                    duration_in_secs=j[1])
            a.save()
            print e

def update_attendance_for_all_users():
    users = User.objects.all()
    for user in User.objects.all():
        populate_attendance_duration_field(user)

def remove_zero_attendance_rows():
    a = AttendanceTracker.objects.all().filter(duration_in_secs=0)
    for i in a:
        i.delete()

def populate_attendance_tracker(user=None):
    """ 
    Purpose is to populate the attendance tracker for user. 
    This is intended to initialize the attendance tracker table
    installed feb 2016. After installation attendance tracker
    objects are automatically created/updated as a side effect of
    an activitylog creation event. see ActivityLog.save() method.
    """

    logs = user.activitylog.all().order_by('timestamp')  # read in ascending order  

    for i in logs:
        try:
            a = AttendanceTracker( 
                user=user, 
                datestamp=i.timestamp, 
                datestr=i.timestamp.astimezone(tz).strftime('%Y-%m-%d'))
            a.save()
            print a.datestamp.strftime('%Y-%m-%d %-I:%M %p %Z'), '==>', a.datestr
        
        except IntegrityError as e:
            pass
        

    print 'Done'

def temp_attendance_report(user=None):
    curr = datetime.now(tz)
    report = user.attendance.all().filter(datestr__startswith=curr.strftime('%Y-%m'))
    import calendar
    days = calendar.monthrange(curr.year, curr.month)
    listing = [None] * days[1]    

    for i in report:
        listing[i.day_tz()-1] = i

    return listing

def update_bookmark_course_context(user_id=None, old_course_id=None, new_course_id=None):
    u = User.objects.get(pk=user_id)
    c = Course.objects.get(pk=new_course_id)
    bks = u.bookmarker.all()
    upd_cnt = 0
    del_cnt = 0
    for i in bks:
        if i.course_context.id == old_course_id:
            try:
                i.course_context = c
                i.save()
                upd_cnt += 1
            except IntegrityError as e:
                print e
                i.delete()
                del_cnt += 1


    print upd_cnt, 'bookmarks updated to course', c.id, c, del_cnt, 'deleted'

def activity_stat_worksheet(u, lesson_obj=None):
    """ Returns a lesson map mapping total number of worksheets and total user completions excludes pretest lessons
    """

    if not lesson_obj:  # get stats for all lessons
        lesson_worksheet_map = {lesson_obj: [lesson_obj.worksheets.all().count(), 0] for lesson_obj in Lesson.objects.all() if lesson_obj.title[-8:] != 'Pretests'}
        for i in u.completed_worksheets.all():
            lesson_worksheet_map[i.completed_worksheet.lesson][1] += 1
    
    else: # get stats by lesson object parameter
        lesson_worksheets = lesson_obj.worksheets.all().count()
        user_completed = u.completed_worksheets.all().filter(completed_worksheet__lesson=lesson_obj).count()
        lesson_worksheet_map = {lesson_obj: [lesson_worksheets, user_completed]}

    return lesson_worksheet_map

def activity_stat_slides(u, lesson_obj=None):
    """ Returns a presentation map mapping total number of presentations and total user views.
    """
    if not lesson_obj:  # get stats for all lessons
        lesson_stack_map = {i.title: [i.slidestacks.all().count(), 0] for i in Lesson.objects.all() if i.title[-8:] != 'Pretests'}
        events = u.activitylog.all().filter(action='access-presentation')
        stackviews = {i.message.rfind('/'): i.message_detail for i in events}
        for i, j in stackviews.items():
            # look up the associated lesson for the event and increment.
            lesson_stack_map[j][1] += 1
    
    else: # get stats by lesson object parameter
        lesson_stacks = lesson_obj.slidestacks.all().count()
        events = u.activitylog.all().filter(message_detail=lesson_obj.title).filter(action='access-presentation')
        # build dict {presentation id: lesson title}
        stackviews = {i.message.rfind('/'): lesson_obj.title for i in events}
        lesson_stack_map = {lesson_obj: [lesson_stacks, len(stackviews)]}

    return lesson_stack_map


