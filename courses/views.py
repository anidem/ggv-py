# courses/views.py

from collections import OrderedDict
from pytz import timezone
import datetime as dt
from datetime import datetime
import calendar
import time
import csv

from django.views.generic import TemplateView, DetailView, UpdateView, CreateView, DeleteView
from django.http import HttpResponse, Http404, HttpResponseRedirect
from django.core.urlresolvers import reverse
from django.core.exceptions import PermissionDenied
from django.contrib.auth.models import User
from django.conf import settings
from django.utils import html, text
from django.shortcuts import redirect, get_object_or_404
from django.contrib.contenttypes.models import ContentType

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from braces.views import LoginRequiredMixin, JSONResponseMixin, AjaxResponseMixin

from core.models import Notification, SiteMessage, BOOKMARK_TYPES, DEACTIVATION_TYPES, SitePage
from core.mixins import AccessRequiredMixin, PrivelegedAccessMixin, RestrictedAccessZoneMixin, CourseContextMixin, GGVUserViewRestrictedAccessMixin
from core.utils import UnicodeWriter, GGVExcelWriter, get_daily_log_times, get_daily_log_times_v2, elapsed_time_per_event
from questions.models import QuestionSet, UserWorksheetStatus
from slidestacks.models import SlideStack
from pretests.models import PretestAccount
from .models import GGVOrganization, Course, CourseGrader
from .forms import CourseUpdateForm, CourseUpdateGraderForm

tz = timezone(settings.TIME_ZONE)

""" Organization Management """

class GgvOrgAdminView(LoginRequiredMixin, DetailView):
    """
        Displays an overview of a ggv organization. This view is intended for users assigned as managers. 
        Information displayed includes: instructor and student list of those assigned to courses, 
        license information, and organization information.

        Visibility: sysadmin, staff, manager(s) assigned to the course
    """
    model = GGVOrganization
    template_name = 'ggvorg_manage.html'
    access_object = None

    def get_context_data(self, **kwargs):
        context = super(GgvOrgAdminView, self).get_context_data(**kwargs)
        org = self.get_object()
        tag = self.request.GET.get('scope', None)
        
        user_licenses_used = org.licenses_in_use(scope=tag)
        
        context['active'] = user_licenses_used['active']
        context['unvalidated'] = user_licenses_used['unvalidated']
        context['num_licensees'] = user_licenses_used['count']
        context['deactivated_users'] = org.deactivated_users()
        # context['courses'] = org.organization_courses.all()
        context['courses'] = user_licenses_used['courses']
        context['licenseinfo'] = user_licenses_used
        context['tag_filter'] = user_licenses_used['tag_filter']

        if self.request.user in org.manager_list(): context['roles'] = ['manage']
        
        # calculate pretest quota info for the entire organization
        pretest_accounts = org.pretest_account_associations.all()
        assigned, purchased = 0, 0
        for i in pretest_accounts:
            assigned = assigned + i.get_pretest_assignments().count()
            purchased = purchased + i.tests_purchased

        context['pretest_stat'] = {
            'purchased': purchased, 
            'used': assigned,
            'available': purchased - assigned
        }
        
        context['pretest_stat']
        # context['is_manager'] = self.request.user.has_perm('manage', course)
        # context['managers'] = course.manager_list()
        # context['instructors'] = instructors
        # context['students'] = students
        # context['deactivated'] = deactivated_students
        # context['unvalidated'] = course.unvalidated_list()

        return context


class GgvOrgUserActivityReportView(LoginRequiredMixin, DetailView):
    """
    Summarizes all student activity in an organization for a specified date (scope). Reports are arranged in
    a single worksheet. This includes all activity for each student on specified day (scope).
    """
    model = GGVOrganization
    template_name = 'ggvorg_user_progress.html'
    access_object = None

    def get(self, request, *args, **kwargs):
        return super(GgvOrgUserActivityReportView, self).get(request, *args, **kwargs)

    def render_to_response(self, context, **response_kwargs):
        scope = self.request.GET.get('scope', '')
        if not scope:
            return super(GgvOrgUserActivityReportView, self).render_to_response(context, **response_kwargs)
        if scope:

            COURSE_INFO_CELLS = ['A1', 'B1', 'C1', 'D1']
            # id, first, last , username, last login, date created, site, organization, deactivation date
            DATA_COLS = [
                ('A1', u'Program Id', 8),
                ('B1', u'Course', 10),
                ('C1', u'First', 10),
                ('D1', u'Last', 10),
                ('E1', u'Username', 30),
                ('F1', u'Date', 12),
                ('G1', u'On site(h:m)', 8),
                ('H1', u'Date/Time', 18),
                ('I1', u'Activity', 15),
                ('J1', u'Subject', 15),
                ('K1', u'Current Score', 5),
            ]

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            daystr = datetime.now().strftime('%Y-%m-%d-%I_%M_%p ')
            orgstr = self.get_object().title
            filename = orgstr + '-' + scope + '-report.xlsx'
            response['Content-Disposition'] = 'attachment; filename=' + filename

            # Openpyxl writer
            writer = Workbook()
            ws = writer.get_active_sheet()
            ws.title = orgstr
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.merge_cells('A1:G1')

            # Write course information row and format
            ws.append([self.get_object().title or '', ' Report date: ' + scope])
            ws.append([])  # Blank row
            for i in COURSE_INFO_CELLS:
                ws[i].font = Font(size=10, name='Arial', bold=True)
                ws[i].alignment = Alignment(wrap_text=True)

            # Write data column header and format
            for col_num in xrange(len(DATA_COLS)):
                offset = col_num+1
                cell = ws.cell(row=3, column=offset)
                cell.value = DATA_COLS[col_num][1]
                cell.font = Font(size=10, name='Arial')
                cell.alignment = Alignment(wrap_text=True)
                # set column width
                ws.column_dimensions[get_column_letter(col_num+1)].width = DATA_COLS[col_num][2]

            # Write data rows
            student_data = []
            courses = self.get_object().organization_courses.all()
            for c in courses:
                for i in c.student_report():                
                    user = User.objects.get(username=i[3])  # get user from username
                    activity_log = get_daily_log_times_v2(user, c)
                    # print scope
                    # get activity log info for scope
                    try:
                        daily = activity_log[scope]
                    except:
                        daily = [0, 0, []]
                    
                    ws.append([i[0], c.title, i[1], i[2], i[3], scope, daily[1]])
                    # compile events
                    for k in daily[2]:
                        ws.append(['', '', '', '', '', '', '', k['event_time'], k['activity'].action, k['activity'].message_detail or ' ', k['score'] or ' '])

            writer.save(response)
            response.set_cookie('fileDownload','true');
            return response
    
    def get_context_data(self, **kwargs):
        context = super(GgvOrgUserActivityReportView, self).get_context_data(**kwargs)   
        return context


""" Course Management """

class CourseUpdateView(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, RestrictedAccessZoneMixin, UpdateView):
    model = Course
    template_name = 'course_settings.html'
    slug_url_kwarg = 'crs_slug'
    form_class = CourseUpdateForm
    access_object = None


class CourseGraderLogView(LoginRequiredMixin, TemplateView):
    template_name = 'course_grader_log.html'
    slug_url_kwarg = 'crs_slug'
    access_object = None

    def get(self, request, *args, **kwargs):
        if request.user.is_staff:
            self.courses = CourseGrader.objects.all()
        else:
            self.courses = request.user.courses_to_grade.all()
            if not self.courses:
                raise PermissionDenied          

        return super(CourseGraderLogView, self).get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(CourseGraderLogView, self).get_context_data(**kwargs)        
        essay_content_id = ContentType.objects.get(model='textquestion').id
        
        course_listing = {}
        for i in self.courses:
            if i.course not in course_listing:
                course_listing[i.course] = []
                students = i.course.student_list()
                for student in students:
                    text_responses = student.question_responses.filter(content_type__id=essay_content_id).filter(modified__gt='2017-5-1')
                    essay_responses = [r for r in text_responses if not r.get_question_object().auto_grade]
                    course_listing[i.course].extend(essay_responses)
        
        context['grading_list'] = course_listing  
        return context


class CourseGraderEditView(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, RestrictedAccessZoneMixin, UpdateView):
    model = CourseGrader
    template_name = 'course_grader_settings.html'
    slug_url_kwarg = 'crs_slug'
    form_class = CourseUpdateGraderForm
    access_object = None

    def get_success_url(self):
        return reverse('manage_course', kwargs={'crs_slug': self.get_object().course.slug})

    def get_context_data(self, **kwargs):
        context = super(CourseGraderEditView, self).get_context_data(**kwargs)        
        form = context['form']
        graders = self.get_object().course.instructor_list() + self.get_object().course.manager_list()
        form.fields['grader'].choices = [(' ','--')] + [(i.id, str(i.first_name + ' ' + i.last_name + ', ' + i.email)) for i in graders]
        form.fields['course'].initial = self.get_object().course
        context['course'] = self.get_object().course
        return context


class CourseGraderCreateView(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, RestrictedAccessZoneMixin, CreateView):
    model = CourseGrader
    template_name = 'course_grader_settings.html'
    slug_url_kwarg = 'crs_slug'
    form_class = CourseUpdateGraderForm
    access_object = None

    def dispatch(self, request, *args, **kwargs):
        self.course = get_object_or_404(Course, slug=kwargs['crs_slug']) 
        return super(CourseGraderCreateView, self).dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse('manage_course', kwargs={'crs_slug': self.kwargs['crs_slug']})

    def get_context_data(self, **kwargs):
        context = super(CourseGraderCreateView, self).get_context_data(**kwargs)
        form = context['form']
        graders = self.course.instructor_list() + self.course.manager_list()
        form.fields['grader'].choices = [(' ','--')] + [(i.id, str(i.first_name + ' ' + i.last_name + ', ' + i.email)) for i in graders]
        form.fields['course'].initial = self.course
        context['course'] = self.course
        return context


class CourseGraderDeleteView(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, RestrictedAccessZoneMixin, DeleteView):
    model = CourseGrader
    template_name = 'course_grader_delete.html'
    slug_url_kwarg = 'crs_slug'
    access_object = None

    def get_success_url(self):
        return reverse('manage_course', kwargs={'crs_slug': self.get_object().course.slug})


class CourseView(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, PrivelegedAccessMixin, DetailView):
    model = Course
    template_name = 'course.html'
    slug_url_kwarg = 'crs_slug'
    access_object = None

    def get_context_data(self, **kwargs):
        context = super(CourseView, self).get_context_data(**kwargs)
        course = self.get_object()
        lessons = course.lesson_list()
        context['eng_lessons'] = lessons.filter(
            lesson__language='eng').order_by('lesson__subject')
        context['span_lessons'] = lessons.filter(
            lesson__language='span').order_by('lesson__subject')

        if self.request.user.is_staff or context['is_instructor'] or context['is_manager']:
            students = course.student_list()
            instructors = course.instructor_list()
            context['instructors'] = instructors
            context['students'] = students

            context['notifications'] = Notification.objects.filter(user_to_notify=self.request.user)

        else:
            context['instructors'] = course.instructor_list()

        context['deactivated'] = course.deactivated_list()
        context['unvalidated'] = course.unvalidated_list()

        # this provides access to the user's full list of courses.
        context['courses'] = [
            Course.objects.get(slug=i) for i in self.request.session['user_courses']]

        try:
            context['site_message'] = SiteMessage.objects.get(url_context=reverse('course', kwargs={'crs_slug': course.slug}))
        except:
            context['site_message'] = None

        context['course_home_view'] = True

        eng_lessons = []
        for i in context['eng_lessons']:
            try:
                plan = SitePage.objects.get(slug=i.lesson.subject+'-ed-plan')
                guide = SitePage.objects.get(slug=i.lesson.subject+'-guide')
                steps = SitePage.objects.get(slug='steps-to-completion')
                eng_lessons.append((i, (plan, guide, steps)))
            except:
                eng_lessons.append((i, (None, None, None)))
                
        span_lessons = []
        for i in context['span_lessons']:
            try:
                plan = SitePage.objects.get(slug=i.lesson.subject+'-ed-plan-span')
                guide = SitePage.objects.get(slug=i.lesson.subject+'-guide-span')
                steps = SitePage.objects.get(slug='steps-to-completion-span')
                span_lessons.append((i, (plan, guide, steps)))
            except:
                span_lessons.append((i, (None, None, None)))


        
        context['english_lessons'] = eng_lessons
        context['spanish_lessons'] = span_lessons
        return context


class CourseManageView(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, RestrictedAccessZoneMixin, PrivelegedAccessMixin, DetailView):
    """
        Displays an overview of student status and activity. Student status is
        either active, deactivated, or not validated (student account is created
        but student has never logged in.)
        The most recent activity for each active student is displayed.
        This view also provides a Course Settings button to allow instructors to modify
        settings that apply to the course.

        Visibility: sysadmin, staff, manager, and instructor
    """
    model = Course
    template_name = 'course_manage.html'
    slug_url_kwarg = 'crs_slug'
    access_object = None

    def get_context_data(self, **kwargs):
        context = super(CourseManageView, self).get_context_data(**kwargs)
        course = self.get_object()
        students = []
        for i in course.student_list():
            try:
                activity = i.activitylog.all()[0]
                students.append((i, {'recent_act': activity.action, 'recent_time': activity.timestamp}, i.ggvuser))
            except:
                pass  # student[i] has no activity on record. Move on, nothing to see here.

        deactivated_students = []        
        for i in course.deactivated_list():
            try:
                activity = i.activitylog.all()[0]
                deactivated_students.append((i, {'recent_act': activity.action, 'recent_time': activity.timestamp}, i.ggvuser))
            except:
                pass  # deactivated_students[i] has no activity on record. Move on, nothing to see here.

        manager_list = course.manager_list()
        instructor_list = course.instructor_list()
        instructors = []        
        for i in instructor_list:
            try:
                activity = i.activitylog.all()[0]
                instructors.append((i, {'recent_act': activity.action, 'recent_time': activity.timestamp}))
            except:
                pass  # instructor[i] has no activity on record. Move on, nothing to see here.

        context['is_manager'] = self.request.user.has_perm('manage', course)
        context['managers'] = manager_list
        context['instructors'] = instructors
        context['students'] = students
        context['deactivated'] = deactivated_students
        context['unvalidated'] = course.unvalidated_list()
        context['account_requests'] = course.account_requests.all()
        context['graders'] = course.assigned_graders.all()
        context['deactivation_types'] = DEACTIVATION_TYPES

        if self.request.user in instructor_list or self.request.user in manager_list or self.request.user.is_staff:
            pretest_account = PretestAccount.objects.filter(ggv_org=course.ggv_organization)
            if pretest_account:
                context['pretest_account'] = pretest_account[0]
        return context


class CourseUserReportView(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, RestrictedAccessZoneMixin, PrivelegedAccessMixin, DetailView):
    """
    Summarizes total time on site and active status for all students in a course. One row for each student written to 
    excel spreadsheet.
    """
    model = Course
    template_name = 'course_user_progress.html'
    slug_url_kwarg = 'crs_slug'
    access_object = None

    def get(self, request, *args, **kwargs):
        return super(CourseUserReportView, self).get(request, *args, **kwargs)

    def render_to_response(self, context, **response_kwargs):
        scope = self.request.GET.get('scope', '')
        if not scope:
            scope = 'all'
        if scope:

            COURSE_INFO_CELLS = ['A1', 'B1', 'C1', 'D1']
            # id, first, last , username, last login, date created, site, organization, deactivation date
            DATA_COLS = [
                ('A1', u'Program Id', 15),
                ('B1', u'First', 20),
                ('C1', u'Last', 20),
                ('D1', u'Username', 30),
                ('E1', u'Last login', 20),
                ('F1', u'Date joined', 20),
                ('G1', u'Course', 15),
                ('H1', u'Organization', 15),
                ('I1', u'On site (hours:mins)', 30),
                ('J1', u'Active?', 15),
            ]

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            daystr = datetime.now().strftime('%Y-%m-%d-%I_%M_%p ')
            coursestr = self.get_object().title
            filename = coursestr + '-' + daystr + '-report.xlsx'
            response['Content-Disposition'] = 'attachment; filename=' + filename

            # Openpyxl writer
            writer = Workbook()
            ws = writer.get_active_sheet()
            ws.title = coursestr

            # Write course information row and format
            ws.append([self.get_object().title or '', ' Report date: ' + daystr])
            ws.append([])  # Blank row
            for i in COURSE_INFO_CELLS:
                ws[i].font = Font(size=18, name='Arial', bold=True)

            # Write data column header and format
            for col_num in xrange(len(DATA_COLS)):
                offset = col_num+1
                cell = ws.cell(row=3, column=offset)
                cell.value = DATA_COLS[col_num][1]
                cell.font = Font(size=14, name='Arial')
                # set column width
                ws.column_dimensions[get_column_letter(col_num+1)].width = DATA_COLS[col_num][2]

            # Write data rows
            for i in self.get_object().student_report():
                ws.append(i)
            
            writer.save(response)
            
            return response

        else:
            return super(CourseUserReportView, self).render_to_response(context, **response_kwargs)
    
    def get_context_data(self, **kwargs):
        context = super(CourseUserReportView, self).get_context_data(**kwargs)     
        return context


class CourseUserActivityReportView(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, RestrictedAccessZoneMixin, PrivelegedAccessMixin, DetailView):
    """
    Summarizes all student activity in a course for a specified date (scope). Reports are arranged in
    a single worksheet. This includes all activity for each student.
    """
    model = Course
    template_name = 'ggvorg_user_progress.html' #'course_user_progress.html'
    slug_url_kwarg = 'crs_slug'
    access_object = None

    def get(self, request, *args, **kwargs):
        return super(CourseUserActivityReportView, self).get(request, *args, **kwargs)

    def render_to_response(self, context, **response_kwargs):
        scope = self.request.GET.get('scope', '')
        if not scope:
            return super(CourseUserActivityReportView, self).render_to_response(context, **response_kwargs)
        if scope:

            COURSE_INFO_CELLS = ['A1', 'B1', 'C1', 'D1']
            # id, first, last , username, last login, date created, site, organization, deactivation date
            DATA_COLS = [
                ('A1', u'Program Id', 10),
                ('B1', u'First', 10),
                ('C1', u'Last', 10),
                ('D1', u'Username', 30),
                ('E1', u'Date', 15),
                ('F1', u'On site(h:m)', 15),
                ('G1', u'Date/Time', 10),
                ('H1', u'Activity', 15),
                ('I1', u'Subject', 15),
                ('J1', u'Current Score', 5),

            ]

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            daystr = datetime.now().strftime('%Y-%m-%d-%I_%M_%p ')
            coursestr = self.get_object().title
            filename = coursestr + '-' + daystr + '-report.xlsx'
            response['Content-Disposition'] = 'attachment; filename=' + filename

            # Openpyxl writer
            writer = Workbook()
            ws = writer.get_active_sheet()
            ws.title = coursestr
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.merge_cells('A1:G1')

            # Write course information row and format
            ws.append([self.get_object().title or '', ' Report date: ' + daystr])
            ws.append([])  # Blank row
            for i in COURSE_INFO_CELLS:
                ws[i].font = Font(size=10, name='Arial', bold=True)
                ws[i].alignment = Alignment(wrap_text=True)

            # Write data column header and format
            for col_num in xrange(len(DATA_COLS)):
                offset = col_num+1
                cell = ws.cell(row=3, column=offset)
                cell.value = DATA_COLS[col_num][1]
                cell.font = Font(size=10, name='Arial')
                cell.alignment = Alignment(wrap_text=True)
                # set column width
                ws.column_dimensions[get_column_letter(col_num+1)].width = DATA_COLS[col_num][2]

            # Write data rows
            for i in self.get_object().student_report():                
                user = User.objects.get(username=i[3])  # get user from username
                activity_log = get_daily_log_times_v2(user, self.get_object())
                
                try:
                    daily = activity_log[scope]
                except:
                    daily = [0, 0, []]
                
                ws.append([i[0], i[1], i[2], i[3], scope, daily[1]])
                
                for k in daily[2]:
                    ws.append(['', '', '', '', '', '', k['event_time'], k['activity'].action, k['activity'].message_detail or ' ', k['score'] or ' '])

            writer.save(response)
            response.set_cookie('fileDownload','true');
            return response

        else:
            return super(CourseUserReportView, self).render_to_response(context, **response_kwargs)
    
    def get_context_data(self, **kwargs):
        context = super(CourseUserActivityReportView, self).get_context_data(**kwargs)
   
        return context


class CourseUserActivityFullReportView(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, RestrictedAccessZoneMixin, PrivelegedAccessMixin, DetailView):
    """
    Summarizes all student activity for each student in a course. Student reports are arranged in separate worksheets.
    """
    model = Course
    template_name = 'course_user_progress.html'
    slug_url_kwarg = 'crs_slug'
    access_object = None

    def get(self, request, *args, **kwargs):
        return super(CourseUserActivityFullReportView, self).get(request, *args, **kwargs)

    def render_to_response(self, context, **response_kwargs):
        USER_INFO_CELLS = ['A1', 'B1', 'C1', 'D1']

        DATA_COLS = [
            ('A1', u'Date', 10),
            ('B1', u'Total Time on Curriculum', 15),
            ('C1', u'Date & Time', 17),
            ('D1', u'Activity', 20),
            ('E1', u'More Details', 30),
            ('F1', u'Subject', 15),
            ('G1', u'Score', 5),
        ]

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        daystr = datetime.now().strftime('%Y-%m-%d-%I_%M_%p ')
        filename = self.get_object().slug + '-' + daystr + 'student-activity-report.xlsx'
        response['Content-Disposition'] = 'attachment; filename=' + filename

        # Openpyxl writer
        writer = Workbook()
        sheet_index = 0
        
        for student in self.get_object().student_report():                
            user = User.objects.get(username=student[3])  # get user from username
            if not user.is_active:
                continue
            userstr = user.last_name + '-' + user.first_name
            
            ws = writer.create_sheet(userstr, sheet_index)
            ws.print_options.gridLines=True
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

            # Write user information row and format
            ws.append([user.ggvuser.program_id or '', user.first_name + ' ' + user.last_name, user.email, ' Report date: ' + daystr])
            ws.append([])  # Blank row
            for i in USER_INFO_CELLS:
                ws[i].font = Font(size=14, name='Arial', bold=True)

            # Write data column header and format
            for col_num in xrange(len(DATA_COLS)):
                offset = col_num+1
                cell = ws.cell(row=3, column=offset)
                cell.value = DATA_COLS[col_num][1]
                cell.font = Font(size=12, name='Arial')
                cell.alignment=Alignment(wrap_text=True)
                # set column width
                ws.column_dimensions[get_column_letter(col_num+1)].width = DATA_COLS[col_num][2]

            activity_log = get_daily_log_times_v2(user, self.get_object())
            # Write data rows
            for i, j in activity_log.items():
                ws.append([j[0].date(), j[1]])
                for k in j[2]:
                    ws.append(['', '', k['event_time'].strftime('%Y-%m-%d %-I:%M %p'), k['activity'].action, html.strip_tags(k['activity'].message), k['activity'].message_detail or ' ', k['score'] or ' '])

            sheet_index += 1
            
            
        writer.save(response)
        return response
    
    def get_context_data(self, **kwargs):
        context = super(CourseUserActivityFullReportView, self).get_context_data(**kwargs)
   
        return context


""" User Activity Management """

class UserManageView(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, RestrictedAccessZoneMixin, PrivelegedAccessMixin, DetailView):
    """
        Displays a detailed or raw dump of user activity from activity log table. Data is displayed
        sequentially ordered with most recent activity listed first.

        If request contains query parameter export=csv|xlsx, then response will be a download of csv
        file containing user activity.

        Visibility: sysadmin, staff, instructor
    """

    model = Course
    template_name = 'course_manage_user.html'
    slug_url_kwarg = 'crs_slug'
    access_object = None

    def get_context_data(self, **kwargs):
        context = super(UserManageView, self).get_context_data(**kwargs)
        course = self.get_object()
        user = User.objects.get(pk=self.kwargs['user'])
        
        context['student_user'] = user
        context['activity_log'] = get_daily_log_timesv2(user, course)

        return context


class UserProgressView(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, GGVUserViewRestrictedAccessMixin, PrivelegedAccessMixin, DetailView):
    """
        Displays progress data for a user/student. This data is filtered here to display
        sequential activity related to a users access and completion of worksheets as well
        as when they view presentations.

        Visibility: sysadmin, staff, instructor, and students (added aug 8 2017)
    """
    model = Course
    template_name = 'course_user_progress.html'
    slug_url_kwarg = 'crs_slug'
    access_object = 'user'
    required_privileges = ['access', 'instructor', 'manage']

    def get(self, request, *args, **kwargs):
        return super(UserProgressView, self).get(request, *args, **kwargs)

    def render_to_response(self, context, **response_kwargs):

        if 'xlsx' in self.request.GET.get('export', ''):

            USER_INFO_CELLS = ['A1', 'B1', 'C1', 'D1']

            DATA_COLS = [
                ('A1', u'Date', 10),
                ('B1', u'Total Time on Curriculum', 15),
                ('C1', u'Date & Time', 17),
                ('D1', u'Activity', 20),
                ('E1', u'More Details', 30),
                ('F1', u'Subject', 15),
                ('G1', u'Current Score', 5),
            ]

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            daystr = datetime.now().strftime('%Y-%m-%d-%I_%M_%p ')
            userstr = context['student_user'].last_name + '-' + context['student_user'].first_name
            filename = userstr + '-' + daystr + '-activity-report.xlsx'
            response['Content-Disposition'] = 'attachment; filename=' + filename

            # Openpyxl writer
            writer = Workbook()
            ws = writer.get_active_sheet()
            ws.title = userstr
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

            # Write user information row and format
            ws.append([context['student_user'].ggvuser.program_id or '', context['student_user'].first_name + ' ' + context['student_user'].last_name, context['student_user'].email, ' Report date: ' + daystr])
            
            # Write user information time on each subject
            for i, j in context['subject_time'].items():
                ws.append([i, str(j[0]) + ' hours' + str(j[1]) + ' mins'])

            ws.append([])  # Blank row

            for i in USER_INFO_CELLS:
                ws[i].font = Font(size=14, name='Arial', bold=True)

            # Write data column header and format
            for col_num in xrange(len(DATA_COLS)):
                offset = col_num+1
                cell = ws.cell(row=10, column=offset)
                cell.value = DATA_COLS[col_num][1]
                cell.font = Font(size=12, name='Arial')
                cell.alignment=Alignment(wrap_text=True)
                # set column width
                ws.column_dimensions[get_column_letter(col_num+1)].width = DATA_COLS[col_num][2]

            # Write data rows for all activity or within date range

            try:
                start = context['start']
                end = context['end']
            except:
                # range not provided
                start, end = None, None
            
            if start and end:  # compile report within date range
                dates = [start + dt.timedelta(days=d) for d in range(0, (end-start).days+1)]
                for i in dates:
                    try:
                        day = context['activity_log'][str(i.date())]
                        ws.append([day[0].date(), day[1]])
                        for k in day[2]:
                            ws.append(['', '', k['event_time'], k['activity'].action, html.strip_tags(k['activity'].message), k['activity'].message_detail or ' ', k['score'] or ' '])
                    except:
                        # 'no activity on current day'
                        pass
            else:  # compile report for all activity
                for i, j in context['activity_log'].items():
                    ws.append([j[0].date(), j[1]])
                    for k in j[2]:
                        ws.append(['', '', k['event_time'], k['activity'].action, html.strip_tags(k['activity'].message), k['activity'].message_detail or ' ', k['score'] or ' '])

            writer.save(response)
            return response

        else:
            return super(UserProgressView, self).render_to_response(context, **response_kwargs)
    
    def get_context_data(self, **kwargs):
        context = super(UserProgressView, self).get_context_data(**kwargs)
        user = User.objects.get(pk=self.kwargs['user'])
        course = self.get_object()
        context['student_user'] = user
        context['activity_log'] = get_daily_log_times_v2(user, course) # 'login', 'logout', 'access-worksheet'

        start = self.request.GET.get('start', '')
        end = self.request.GET.get('end', '')
        if start and end:  # compile subject times within date range
            context['start'] = datetime.strptime(start, "%Y-%m-%d")
            context['end'] = datetime.strptime(end, "%Y-%m-%d")
            
            subject_time = elapsed_time_per_event(user, context['start'], context['end'])[1]
            
        else:
            subject_time = elapsed_time_per_event(user)[1]
        
        for i, j in subject_time.items():
            subject_time[i] = j/3600, (j%3600)/60
        
        context['subject_time'] = subject_time

        if 'completed' in self.request.GET.get('filter', ''):
            context['filter'] = 'completed'

        opts = dict(BOOKMARK_TYPES)
        for i, j in opts.items():
            if user.ggvuser.language_pref == 'spanish':
                opts[i] = j.split(',')[1]
            else:
                opts[i] = j.split(',')[0]

        context['bookmark_type_opts'] = opts  
        

        return context

class UserProgressViewDateSelector(UserProgressView):
    model = Course
    template_name = 'course_user_progress_custom.html'
    slug_url_kwarg = 'crs_slug'
    access_object = 'user'
    required_privileges = ['access', 'instructor', 'manage']

    def get_context_data(self, **kwargs):
        context = super(UserProgressViewDateSelector, self).get_context_data(**kwargs)
        # user = User.objects.get(pk=self.kwargs['user'])
        # context['student_user'] = user
        return context


""" Course Message Management """

class CourseMessageAddView(LoginRequiredMixin, CourseContextMixin, RestrictedAccessZoneMixin, CreateView):
    model = SiteMessage
    template_name = 'ggv_create_page_msg.html'
    course = None
    fields = ['message', 'url_context', 'show']

    def dispatch(self, request, *args, **kwargs):
        try:
            self.course = Course.objects.get(slug=kwargs['crs_slug'])
        except:
            self.course = None

        return super(CourseMessageAddView, self).dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse('course', kwargs={'crs_slug': self.course.slug})

    def get_initial(self):
        self.initial = {'url_context':  reverse('course', kwargs={'crs_slug': self.course.slug})}
        return self.initial


class CourseMessageUpdateView(LoginRequiredMixin, CourseContextMixin, RestrictedAccessZoneMixin, UpdateView):
    model = SiteMessage
    template_name = 'ggv_update_page_msg.html'
    course = None
    fields = ['message', 'show']

    def dispatch(self, request, *args, **kwargs):
        try:
            self.course = Course.objects.get(slug=kwargs['crs_slug'])
        except:
            self.course = None

        return super(CourseMessageUpdateView, self).dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse('course', kwargs={'crs_slug': self.course.slug})


class CourseMessageDeleteView(LoginRequiredMixin, CourseContextMixin, RestrictedAccessZoneMixin, DeleteView):
    model = SiteMessage
    template_name = 'ggv_delete_page_msg.html'
    course = None
    fields = ['message', 'show']

    def dispatch(self, request, *args, **kwargs):
        try:
            self.course = Course.objects.get(slug=kwargs['crs_slug'])
        except:
            self.course = None

        return super(CourseMessageDeleteView, self).dispatch(request, *args, **kwargs)

    def get_success_url(self):
        return reverse('course', kwargs={'crs_slug': self.course.slug})


""" Course Attendance Pages """

class CourseAttendanceMonthView(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, RestrictedAccessZoneMixin, PrivelegedAccessMixin, DetailView):
    model = Course
    template_name = 'course_attendance_current.html'
    slug_url_kwarg = 'crs_slug'
    access_object = None
    date_range = None

    def get(self, request, *args, **kwargs):
        try:
            self.date_range = datetime(int(kwargs['year']), int(kwargs['month']), 1)
        except:
            self.date_range = datetime.now(tz)

        return super(CourseAttendanceMonthView, self).get(request, *args, **kwargs)        

    def get_context_data(self, **kwargs):
        context = super(CourseAttendanceMonthView, self).get_context_data(**kwargs)
        course = self.get_object()       
        months = []
        month_list = []

        students = course.student_list()

        oldest_login = datetime.now().date()
        for student in students:
            if oldest_login > student.date_joined.date():
                oldest_login = student.date_joined.date()

        curr = (datetime.now().date().year, datetime.now().date().month)
        
        if oldest_login.month == 1:
            stop = (oldest_login.year-1, 12)
        else: 
            stop = (oldest_login.year, oldest_login.month-1)

        while curr != stop:
            listings_by_month = []
            month_list.append(curr)

            # date object for current year and month
            listings_by_month.append(datetime(curr[0], curr[1], 1))

            # calendar information for current year and month
            mon = calendar.Calendar()
            day_list = []
            for i in mon.itermonthdays2(curr[0], curr[1]):
                if i[0] > 0:
                    day_list.append( (calendar.day_abbr[i[1]], i[0]) )
            listings_by_month.append(day_list)

            # build and attach attendance sheet for current year and month
            for student in students:
                user_attendance = student.ggvuser.attendance_by_month(curr[0], curr[1])
                listings_by_month.append({ student : user_attendance })

            months.append(listings_by_month)

            if (curr[1]-1) % 12 == 0:
                curr = (curr[0]-1, 12)
            else:
                curr = (curr[0],curr[1]-1)




        # deprecated ==> days_in_month = range(1, calendar.monthrange(self.date_range.year, self.date_range.month)[1]+1)



        context['attendance_sheet'] = months
        context['month_year'] = self.date_range.strftime('%B %Y')
        context['date_display'] = self.date_range
        context['months'] = month_list
        return context


class CourseAttendanceUserView(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, RestrictedAccessZoneMixin, PrivelegedAccessMixin, DetailView):
    model = Course
    template_name = 'course_attendance_user.html'
    slug_url_kwarg = 'crs_slug'
    access_object = None
    student = None

    def get(self, request, *args, **kwargs):
        try:
            self.student = User.objects.get(pk=kwargs['user'])
        except:
            raise Http404("Student records not found.")

        return super(CourseAttendanceUserView, self).get(request, *args, **kwargs)        

    def get_context_data(self, **kwargs):
        context = super(CourseAttendanceUserView, self).get_context_data(**kwargs)  
        context['attendance_sheet'] = self.student.ggvuser.attendance_full_listing()
        context['student'] = self.student
        
        return context


""" Analytics """

# class UserElapsedTimePerActivity(LoginRequiredMixin, CourseContextMixin, AccessRequiredMixin, RestrictedAccessZoneMixin, PrivelegedAccessMixin, DetailView)

#     model = Course
#     template_name = 'course_user_activity_time.html'
#     slug_url_kwarg = 'crs_slug'
#     access_object = None

#     def get_context_data(self, **kwargs):
#         context = super(UserElapsedTimePerActivity, self).get_context_data(**kwargs)
#         user = User.objects.get(pk=self.kwargs['user'])
#         context['student_user'] = user
#         context['activity_log'] = elapsed_time_per_event(user) # 'login', 'logout', 'access-worksheet'       
#         return context
